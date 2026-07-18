"""
Stage 7 -- Excel Report Generation
Dataset Type: Synthetic
Creates comprehensive workbook with all results.
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings("ignore")

from src.config import (DATA_CLEAN_DIR, FORECASTS_DIR, REPORTS_DIR,
                         LOCATION_ORDER, WATERMARK)

# Style constants
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="E2E8F0", size=11)
DATA_FONT = Font(name="Calibri", size=10, color="333333")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1E293B")
SUBTITLE_FONT = Font(name="Calibri", italic=True, size=10, color="64748B")
BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
SYNTH_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
SYNTH_FONT = Font(name="Calibri", bold=True, size=10, color="92400E")

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

def add_synth_notice(ws, row):
    ws.cell(row=row, column=1, value="Dataset Type: Synthetic").font = SYNTH_FONT
    ws.cell(row=row, column=1).fill = SYNTH_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

def write_df_to_sheet(ws, df, start_row=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == start_row:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.font = DATA_FONT
                cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
    
    # Auto-width
    for col_cells in ws.columns:
        valid_cells = [c for c in col_cells if hasattr(c, 'column_letter')]
        if not valid_cells:
            continue
        column_letter = valid_cells[0].column_letter
        max_length = 0
        for cell in valid_cells:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 4, 30)
    
    return start_row + len(df) + 1

def main():
    print("=" * 60)
    print(f"STAGE 7 -- EXCEL REPORT GENERATION  |  {WATERMARK}")
    print("=" * 60)
    
    wb = Workbook()
    
    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value="Multi-Hospital Admission Forecasting Study").font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws.cell(row=2, column=1, value="A Time-Series and Machine Learning Study Across Six Locations").font = SUBTITLE_FONT
    ws.merge_cells("A2:F2")
    add_synth_notice(ws, 4)
    
    ws.cell(row=6, column=1, value="Project Overview").font = Font(bold=True, size=12, name="Calibri")
    info = [
        ("Analysis Period", "2022-01 to 2026-06"),
        ("Forecast Horizon", "Jul 2026 - Dec 2026 (6 months)"),
        ("Locations", ", ".join(LOCATION_ORDER)),
        ("Time-Series Models", "SARIMA, ETS (Holt-Winters), Prophet"),
        ("ML Models", "XGBoost, Random Forest, Linear Regression (Ridge)"),
        ("Evaluation Metrics", "MAE, RMSE, MAPE, R-squared"),
        ("Data Source", "Synthetic datasets (no real patient data)"),
        ("Important Notice", "Results are based on synthetic data and should not be "
         "interpreted as identical to original data analysis results."),
    ]
    for i, (key, val) in enumerate(info):
        ws.cell(row=8 + i, column=1, value=key).font = Font(bold=True, name="Calibri", size=10)
        ws.cell(row=8 + i, column=2, value=val).font = DATA_FONT
        ws.merge_cells(start_row=8+i, start_column=2, end_row=8+i, end_column=6)
    
    print("  [OK] Summary sheet")
    
    # ── Sheet 2: EDA Highlights ──
    ws2 = wb.create_sheet("EDA Highlights")
    add_synth_notice(ws2, 1)
    ws2.cell(row=3, column=1, value="Admission Statistics by Location").font = Font(bold=True, size=12, name="Calibri")
    
    monthly = pd.read_csv(os.path.join(DATA_CLEAN_DIR, "monthly_admissions.csv"), parse_dates=["Month"])
    daily = pd.read_csv(os.path.join(DATA_CLEAN_DIR, "daily_admissions.csv"), parse_dates=["Admission Date"])
    
    eda_rows = []
    for loc in LOCATION_ORDER:
        m = monthly[monthly["Location"] == loc]["Admissions"]
        d = daily[daily["Location"] == loc]["Admissions"]
        eda_rows.append({
            "Location": loc,
            "Total Admissions": int(d.sum()),
            "Date Range Start": str(daily[daily["Location"] == loc]["Admission Date"].min().date()),
            "Date Range End": str(daily[daily["Location"] == loc]["Admission Date"].max().date()),
            "Months of Data": int(len(m)),
            "Avg Monthly": round(m.mean(), 1),
            "Std Monthly": round(m.std(), 1),
            "Min Monthly": int(m.min()),
            "Max Monthly": int(m.max()),
            "Avg Daily": round(d.mean(), 1),
        })
    eda_df = pd.DataFrame(eda_rows)
    write_df_to_sheet(ws2, eda_df, start_row=5)
    print("  [OK] EDA Highlights sheet")
    
    # ── Sheet 3: Model Performance ──
    ws3 = wb.create_sheet("Model Performance")
    add_synth_notice(ws3, 1)
    ws3.cell(row=3, column=1, value="Model Evaluation Metrics (Test Period: Jan-Jun 2026)").font = Font(bold=True, size=12, name="Calibri")
    
    metrics_path = os.path.join(FORECASTS_DIR, "model_metrics.csv")
    if os.path.exists(metrics_path):
        metrics = pd.read_csv(metrics_path)
        metrics = metrics[["Location", "Model", "MAE", "RMSE", "MAPE", "R2"]]
        # Round numeric columns
        for col in ["MAE", "RMSE", "MAPE", "R2"]:
            metrics[col] = metrics[col].round(2)
        write_df_to_sheet(ws3, metrics, start_row=5)
    print("  [OK] Model Performance sheet")
    
    # ── Sheet 4: Forecasts ──
    ws4 = wb.create_sheet("Forecasts")
    add_synth_notice(ws4, 1)
    ws4.cell(row=3, column=1, value="Monthly Forecasts: Jul-Dec 2026").font = Font(bold=True, size=12, name="Calibri")
    
    forecast_path = os.path.join(FORECASTS_DIR, "future_forecasts.csv")
    if os.path.exists(forecast_path):
        forecasts = pd.read_csv(forecast_path)
        # Round numeric columns
        num_cols = forecasts.select_dtypes(include=[np.number]).columns
        for col in num_cols:
            forecasts[col] = forecasts[col].round(0)
        write_df_to_sheet(ws4, forecasts, start_row=5)
    print("  [OK] Forecasts sheet")
    
    # ── Sheet 5: Department Analysis ──
    ws5 = wb.create_sheet("Department Analysis")
    add_synth_notice(ws5, 1)
    ws5.cell(row=3, column=1, value="Top Departments by Admission Volume").font = Font(bold=True, size=12, name="Calibri")
    
    dept = pd.read_csv(os.path.join(DATA_CLEAN_DIR, "department_monthly.csv"))
    dept_summary = (dept.groupby(["Location", "Admitting Department"])["Admissions"]
                    .sum().reset_index()
                    .sort_values(["Location", "Admissions"], ascending=[True, False]))
    
    # Keep top 10 departments per location
    top_depts = dept_summary.groupby("Location").head(10).reset_index(drop=True)
    write_df_to_sheet(ws5, top_depts, start_row=5)
    print("  [OK] Department Analysis sheet")
    
    # ── Sheet 6: Best Models ──
    ws6 = wb.create_sheet("Best Models")
    add_synth_notice(ws6, 1)
    ws6.cell(row=3, column=1, value="Best Performing Model per Location").font = Font(bold=True, size=12, name="Calibri")
    
    best_path = os.path.join(FORECASTS_DIR, "best_models.csv")
    if os.path.exists(best_path):
        best = pd.read_csv(best_path)
        write_df_to_sheet(ws6, best, start_row=5)
    print("  [OK] Best Models sheet")
    
    # Save
    output_path = os.path.join(REPORTS_DIR, "Hospital_Admission_Forecasting_Report.xlsx")
    wb.save(output_path)
    print(f"\n[OK] Stage 7 complete -- report saved to {output_path}")

if __name__ == "__main__":
    main()
